from __future__ import annotations

import argparse
import json
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
DEFAULT_SOURCE = DATA_DIR / "imported_course_info.xlsx"
DEFAULT_TARGET = DATA_DIR / "course_assignments.json"
SKIP_TEACHER_MARKS = ("具体上课安排", "个人课表", "待定", "未定", "无")


def load_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return re.sub(r"\s+", " ", text)


def parse_credits(value: Any) -> int | float:
    text = clean(value)
    if not text:
        return 0
    try:
        number = float(text)
    except ValueError:
        match = re.search(r"\d+(?:\.\d+)?", text)
        if not match:
            return 0
        number = float(match.group(0))
    return int(number) if number.is_integer() else number


def split_people(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    parts = re.split(r"[;；、,，\n]+", text)
    people = []
    for part in parts:
        name = clean(part)
        if not name:
            continue
        if any(mark in name for mark in SKIP_TEACHER_MARKS):
            continue
        people.append(name)
    return list(dict.fromkeys(people))


def split_majors(value: Any) -> list[str]:
    text = clean(value)
    if not text:
        return []
    parts = re.split(r"[;；，\n]+", text)
    return [item for item in (clean(part) for part in parts) if item]


def teacher_college_index() -> dict[str, str]:
    index: dict[str, str] = {}
    roster = load_json(DATA_DIR / "teacher_roster.json", {"teachers": []})
    for row in roster.get("teachers", []):
        name = clean(row.get("name"))
        college = clean(row.get("college"))
        if name and college and name not in index:
            index[name] = college

    profiles = load_json(DATA_DIR / "faculty_profiles.json", {"teachers": []})
    for row in profiles.get("teachers", []):
        name = clean(row.get("name"))
        colleges = [clean(item) for item in row.get("colleges", []) if clean(item)]
        if name and colleges and name not in index:
            index[name] = colleges[0]
    return index


def normalize_college(value: Any, teacher: str, index: dict[str, str]) -> str:
    text = clean(value)
    if not text or "未在官方名单" in text:
        return index.get(teacher, "")
    return text


def import_assignments(source: Path, target: Path) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    teacher_colleges = teacher_college_index()
    assignments: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [clean(value) for value in next(rows)]
        except StopIteration:
            continue
        header_map = {name: index for index, name in enumerate(headers)}

        def pick(row: tuple[Any, ...], *names: str) -> Any:
            for name in names:
                if name in header_map and header_map[name] < len(row):
                    return row[header_map[name]]
            return ""

        for row in rows:
            course = clean(pick(row, "课程名", "课程名称", "课程"))
            if not course:
                continue
            credits = parse_credits(pick(row, "学分"))
            majors = split_majors(pick(row, "专业"))
            campus = clean(pick(row, "校区"))
            term = clean(pick(row, "学期"))
            raw_college = pick(row, "学院", "开课学院")
            for teacher in split_people(pick(row, "授课老师", "任课教师", "教师", "老师")):
                college = normalize_college(raw_college, teacher, teacher_colleges)
                key = (teacher, course, credits, tuple(majors), campus, term)
                if key in seen:
                    continue
                seen.add(key)
                assignments.append(
                    {
                        "teacher": teacher,
                        "course": course,
                        "credits": credits,
                        "majors": majors,
                        "college": college,
                        "campus": campus,
                        "term": term,
                        "source_sheet": sheet.title,
                    }
                )

    payload = {
        "source": {
            "name": "课程信息汇总表",
            "file": str(source),
            "imported_at": time.strftime("%Y-%m-%d %H:%M:%S"),
            "row_count": len(assignments),
        },
        "assignments": sorted(assignments, key=lambda row: (row["teacher"], row["course"], row["term"])),
    }
    with target.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return payload["source"]


def main() -> None:
    parser = argparse.ArgumentParser(description="Import course-teacher-major-credit assignments.")
    parser.add_argument("source", nargs="?", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    args = parser.parse_args()
    result = import_assignments(args.source, args.target)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
