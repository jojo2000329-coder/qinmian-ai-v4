from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CATEGORY_COLUMNS = {
    "通识教育必修": 3,
    "通识教育选修": 4,
    "专业基础课": 5,
    "专业核心课": 6,
    "专业选修课": 7,
    "专业实践": 8,
    "社会实践": 9,
}


def cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def minimum_credit(value: Any) -> float | int | None:
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return None
    match = re.search(r"\d+(?:\.\d+)?", value)
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


def section_markers(sheet) -> list[dict[str, Any]]:
    labels = {"境内生": "domestic", "境外生": "international"}
    markers = []
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            label = cell_value(sheet.cell(row, col).value)
            if label in labels:
                markers.append({"row": row, "col": col, "label": label, "id": labels[label]})
    return markers


def read_section(
    sheet,
    marker: dict[str, Any],
    markers: list[dict[str, Any]],
) -> dict[str, Any]:
    start_col = marker["col"]
    start_row = marker["row"] + 2
    later_same_column = [
        item["row"]
        for item in markers
        if item["col"] == start_col and item["row"] > marker["row"]
    ]
    end_row = min(later_same_column) - 1 if later_same_column else sheet.max_row
    records = []
    for row in range(start_row, end_row + 1):
        college = cell_value(sheet.cell(row, start_col).value)
        major = cell_value(sheet.cell(row, start_col + 1).value)
        total = cell_value(sheet.cell(row, start_col + 2).value)
        if not college or not major or total is None:
            continue

        categories = {
            name: cell_value(sheet.cell(row, start_col + offset).value)
            for name, offset in CATEGORY_COLUMNS.items()
        }
        minimums = {
            name: minimum_credit(value)
            for name, value in categories.items()
        }
        numeric_minimums = [value for value in minimums.values() if isinstance(value, (int, float))]
        can_reconcile = len(numeric_minimums) == len(minimums)
        category_total = sum(numeric_minimums) if can_reconcile else None
        difference = category_total - total if can_reconcile and isinstance(total, (int, float)) else None
        records.append(
            {
                "college": college,
                "major": major,
                "graduation_total": total,
                "categories": categories,
                "category_minimums": minimums,
                "validation": {
                    "category_total": category_total,
                    "difference_from_graduation_total": difference,
                    "matches_graduation_total": abs(difference) < 0.001 if difference is not None else None,
                },
                "source_row": row,
            }
        )

    return {
        "id": marker["id"],
        "label": marker["label"],
        "records": records,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="导入境内生/境外生毕业学分结构表")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = load_workbook(args.source, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    markers = section_markers(sheet)
    sections = {marker["id"]: read_section(sheet, marker, markers) for marker in markers}
    if not {"domestic", "international"}.issubset(sections):
        raise ValueError("未能同时识别“境内生”和“境外生”数据区块")
    document = {
        "source": {
            "name": args.source.name,
            "type": "用户提供的真实毕业学分结构表",
            "sheet": sheet.title,
            "imported_at": date.today().isoformat(),
            "notes": [
                "数据分为境内生和境外生两套要求。",
                "总学分与分类学分保留原表显示值。",
                "带说明或范围的分类学分同时提取最低数值，供学分预警使用。",
                "程序会校验分类最低学分合计与毕业总学分是否一致，并保留源表中的异常。",
            ],
        },
        "student_types": sections,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(document, ensure_ascii=False, indent=2), encoding="utf-8")
    counts = {key: len(value["records"]) for key, value in document["student_types"].items()}
    print(json.dumps(counts, ensure_ascii=False))


if __name__ == "__main__":
    main()
